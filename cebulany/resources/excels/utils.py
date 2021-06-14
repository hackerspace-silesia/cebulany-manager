from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from flask import send_file


def send_excel(workbook: Workbook, download_name):
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)

        return send_file(
            tmp.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=download_name,
        )


def setup_styles(workbook):
    font_args = dict(name='Calibri', size=10)
    header = NamedStyle(name="header")
    header.font = Font(bold=True, name='Calibri', size=12)
    header.alignment = Alignment(horizontal='center')
    workbook.add_named_style(header)

    left_header = NamedStyle(name="left_header")
    left_header.font = Font(bold=True, **font_args)
    workbook.add_named_style(left_header)

    inactive = NamedStyle(name="inactive")
    inactive.fill = PatternFill(start_color='D6D6D6', fill_type='solid')
    inactive.font = Font(bold=True, **font_args)
    workbook.add_named_style(inactive)

    bad = NamedStyle(name="bad")
    bad.fill = PatternFill(start_color='f1b0b7', fill_type='solid')
    bad.font = Font(bold=True, **font_args)
    workbook.add_named_style(bad)

    ok = NamedStyle(name="ok")
    ok.font = Font(**font_args)
    workbook.add_named_style(ok)
