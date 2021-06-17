from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from cebulany.auth import token_required
from cebulany.queries.member import MemberQuery
from cebulany.queries.paid_month import PaidMonthQuery
from cebulany.resources.excels.utils import send_excel, setup_styles, add_cell
from cebulany.resources.excels.blueprint import excel_page, URL_PREFIX


@excel_page.route(URL_PREFIX + '/table/<int:start_year>-<int:end_year>/<int:payment_type_id>')
@token_required
def excel_paid_month(start_year: int, end_year: int, payment_type_id: int):
    members = MemberQuery.get_list_query(order="table")
    all_paid_months = {
        paid_month['member_id']: paid_month['months']
        for paid_month in PaidMonthQuery.get_aggregated_payments(
            payment_type_id=payment_type_id,
            start_year=start_year,
            end_year=end_year,
        )
    }

    workbook = gen_workbook(
        year_span=(start_year, end_year),
        members=members,
        all_paid_months=all_paid_months,
    )

    download_name = f'members-{start_year}-{end_year}.xlsx'
    return send_excel(workbook, download_name)


def gen_workbook(year_span, members, all_paid_months):
    workbook = Workbook()
    setup_styles(workbook)
    sheet = workbook.active
    sheet.title = 'Members'
    add_header(sheet, year_span)
    add_content(sheet, year_span, members, all_paid_months)

    return workbook


def add_header(sheet, year_span):
    first_row = [add_cell(sheet)]
    second_row = [add_cell(sheet)]
    years = list(range(year_span[0], year_span[1] + 1))

    for year in years:
        cell = add_cell(sheet, str(year), 'header')
        first_row += [cell] + [add_cell(sheet)] * 11

        for month in range(1, 12 + 1):
            cell = add_cell(sheet, f'{month:02d}', 'header')
            second_row.append(cell)

    sheet.append(first_row)
    sheet.append(second_row)
    sheet.column_dimensions[get_column_letter(1)].auto_size = True

    for i in range(len(years)):
        i *= 12
        sheet.merge_cells(
            start_row=1,
            end_row=1,
            start_column=i + 1 + 1,
            end_column=i + 12 + 1,
        )


def add_content(sheet, year_span, members, all_paid_months):
    months = list(gen_months(*year_span))
    for member in members:
        paid_months = all_paid_months.get(member.id, {})

        today = datetime.utcnow().strftime('%Y-%m')
        join_date = member.join_date.strftime('%Y-%m')
        is_active = member.is_active
        header = add_cell(sheet, member.name, 'left_header')
        row = [header]
        for month in months:
            month_info = paid_months.get(month)
            cell = gen_cell(sheet, month_info, month, join_date, is_active, today)
            row.append(cell)
        sheet.append(row)


def gen_cell(sheet, month_info, month, join_date, is_active, today):
    cell = add_cell(sheet)
    if month_info is not None:
        cell.value = month_info['sum']
    elif join_date <= month and today >= month:
            cell.style = 'bad'
    if not is_active:
        cell.style = 'inactive'
    return cell


def gen_months(start_year, end_year):
    for year in range(start_year, end_year + 1):
        for month in range(1, 12 + 1):
            yield f'{year:04d}-{month:02d}'
