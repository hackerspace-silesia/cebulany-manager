from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from cebulany.auth import token_required
from cebulany.models import Budget, PaymentType
from cebulany.queries.payment_summary import PaymentSummaryQuery
from cebulany.resources.excels.blueprint import URL_PREFIX, excel_page
from cebulany.resources.excels.utils import send_excel, setup_styles, add_cell


@excel_page.route(URL_PREFIX + '/summary/<int:year>')
@token_required
def excel_payment_summary(year: int):
    payments = {
        (o['budget_id'], o['payment_type_id'], o['is_positive']): o['cost']
        for o in PaymentSummaryQuery.get_payment_data(year)
    }
    budget_ids = set(id for id, _, _ in payments.keys())
    payment_type_ids = set(id for _, id, _ in payments.keys())
    workbook = gen_workbook(
        payments={
            (o['budget_id'], o['payment_type_id'], o['is_positive']): o['cost']
            for o in PaymentSummaryQuery.get_payment_data(year)
        },
        budgets=list(
            Budget.query.filter(Budget.id.in_(budget_ids))
        ),
        payment_types=list(
            PaymentType.query.filter(PaymentType.id.in_(payment_type_ids))
        ),
        outstanding_cost=PaymentSummaryQuery.get_outstanding_cost(year),
        balances=PaymentSummaryQuery.get_balances(year),
    )

    download_name = f'summary-{year}.xlsx'
    return send_excel(workbook, download_name)


def gen_workbook(payments, budgets, payment_types, outstanding_cost, balances):
    workbook = Workbook()
    setup_styles(workbook)
    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'
    balances_sheet = workbook.create_sheet(title='Balances')
    add_budget_header(summary_sheet, budgets)
    add_summary_content(summary_sheet, payments, budgets, payment_types)
    add_balances(balances_sheet, balances, outstanding_cost)

    return workbook


def add_budget_header(sheet, budgets):
    first_row = [add_cell(sheet, 'Budżet →')]
    second_row = [add_cell(sheet, 'Typ ↓')]

    for col, budget in enumerate(budgets):
        sheet.column_dimensions[get_column_letter(col * 2 + 2)].width = 15.0
        sheet.column_dimensions[get_column_letter(col * 2 + 3)].width = 15.0
        cell = add_cell(sheet, budget.name, 'header')
        cell.fill = PatternFill(start_color=budget.color, fill_type='solid')
        first_row += [cell, add_cell(sheet)]

        pos_cell = add_cell(sheet, 'przychód')
        pos_cell.fill = PatternFill(start_color='29AB87', fill_type='solid')
        neg_cell = add_cell(sheet, 'koszt')
        neg_cell.fill = PatternFill(start_color='FE4C40', fill_type='solid')
        second_row += [pos_cell, neg_cell]

    sheet.append(first_row)
    sheet.append(second_row)
    sheet.column_dimensions['A'].width = 30.0
    sheet.freeze_panes = 'B3'

    for i in range(len(budgets)):
        i *= 2
        sheet.merge_cells(
            start_row=1,
            end_row=1,
            start_column=i + 1 + 1,
            end_column=i + 2 + 1,
        )


def add_summary_content(sheet, payments, budgets, payment_types):
    for payment_type in payment_types:
        payment_cell = add_cell(sheet, payment_type.name, 'header')
        payment_cell.fill = PatternFill(start_color=payment_type.color, fill_type='solid')
        row = [payment_cell]
        for budget in budgets:
            for is_positive in (True, False):
                key = budget.id, payment_type.id, is_positive
                payment = payments.get(key)
                row.append(gen_summary_cell(sheet, payment, is_positive))
        sheet.append(row)


def gen_summary_cell(sheet, payment, is_positive):
    if not payment:
        return add_cell(sheet)
    if not is_positive:
        payment *= -1
    return add_cell(sheet, value=payment)


def add_balances(sheet, balances, outstanding_cost):
    sheet.column_dimensions[get_column_letter(1)].auto_size = True
    add_balance_col(sheet, 'Nierozliczone', outstanding_cost)
    add_balance_col(sheet, 'Początek roku')
    add_balance_col(sheet, 'Aktywa obrotowe', balances['curr_start_year'])
    add_balance_col(sheet, 'Zysk/Strata z lat ubiegłych', balances['diff_prev_start_year'])
    add_balance_col(sheet, 'Zysk/Strata netto', balances['diff_start_year'])
    add_balance_col(sheet, 'Koniec roku')
    add_balance_col(sheet, 'Aktywa obrotowe', balances['curr_end_year'])
    add_balance_col(sheet, 'Zysk/Strata z lat ubiegłych', balances['diff_prev_end_year'])
    add_balance_col(sheet, 'Zysk/Strata netto', balances['diff_end_year'])


def add_balance_col(sheet, header, value=None):
    header_cell = add_cell(sheet, header, 'left_header')

    if value is None:
        sheet.append([header_cell])
        return

    value_cell = add_cell(sheet, value=str(value))
    sheet.append([header_cell, value_cell])
