#! coding: utf-8
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from itertools import accumulate
from decimal import Decimal
from os import environ
from typing import TypedDict

from flask import Blueprint, render_template
from sqlalchemy import func as sql_func, cast, Integer

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, NamedStyle
from openpyxl.chart import PieChart, Reference, LineChart, AreaChart
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.axis import ChartLines

from cebulany.models import db, Transaction, Payment, Budget, PaymentType

import cebulany.resources.excels
import cebulany.resources.excels.paid_month
from cebulany.resources.excels.utils import setup_styles

URL_PREFIX = environ.get('CEBULANY_APP_URL_PREFIX', '')


report_page = Blueprint('report_page', 'report', template_folder='../templates')
month_field = cast(sql_func.extract('month', Transaction.date), Integer)
year_field = cast(sql_func.extract('year', Transaction.date), Integer)


class PieDict(TypedDict):
    vals: list[float]
    labels: list[str]


class ReportMonth(object):

    def __init__(self, year: int, month: int, total: Decimal):
        self.year = year
        self.month = month
        self.rows: list["Row"] = []
        self.graphs: dict[str, PieDict] = {}
        self.row_groups: dict[str, list["Row"]] = {}

        positive_graphs: list["Row"] = []
        negative_graphs: list["Row"] = []

        rest_cost = 0

        types = db.session.query(PaymentType).order_by(PaymentType.name)
        for payment_type in types:
            if payment_type.show_details_in_report:
                cost, rows, graphs = self.get_payments_with_details(payment_type)
            else:
                cost, rows, graphs = self.get_total_payments(payment_type)

            rest_cost += cost
            positive_graphs += [payment for payment in graphs if payment.value > 0]
            negative_graphs += [payment for payment in graphs if payment.value < 0]
            self.rows += rows

        self.add_graph('positive', positive_graphs)
        self.add_graph('negative', negative_graphs, multiple=-1)

        others: list["Row"] = []
        diff = total - rest_cost
        if abs(diff) > Decimal('0.01'):
            others.append(Money(u'NIE ROZLICZONE', diff))
        others.append(Money(u'RAZEM', total))

        self.add_graph('others', others)
        self.rows += others

    def get_total_payments(self, payment_type):
        query = (
            db.session
            .query(sql_func.sum(Payment.cost), sql_func.count(Payment.id))
            .join(Payment.transaction)
            .filter(Payment.payment_type_id == payment_type.id)
        )
        query = self.filterize_query(query)
        name = payment_type.name.upper()
        query_row = query.first()
        cost, count = query_row
        if cost is None:
            return 0, [], []
        payment = Money(name, cost)
        rows = [payment]

        # broken - todo fix
        #if payment_type.show_count_in_report:
        #    rows.append(Row(u'{}: {}'.format(name, u'ILOŚĆ'), count))

        return cost, rows, [payment]

    def get_payments_with_details(self, payment_type):
        query = (
            db.session
            .query(sql_func.upper(Payment.name), sql_func.sum(Payment.cost))
            .join(Payment.transaction)
            .filter(Payment.payment_type_id == payment_type.id)
            .group_by(sql_func.upper(Payment.name))
        )
        query = self.filterize_query(query)
        payment_type_name = payment_type.name.upper()

        payments = [
            Money(f'{payment_type_name}: {name}', cost)
            for name, cost in query.all()
        ]

        if len(payments) == 0:
            return 0, [], []
        total = sum((Decimal(obj.value) for obj in payments), Decimal('0.00'))
        rows = payments + [Money(payment_type_name + u': SUMA', total)]

        return total, rows, payments

    def filterize_query(self, query):
        if self.month is None:
            return query.filter(year_field == self.year)
        else:
            return query.filter(
                year_field == self.year,
                month_field == self.month,
            )

    def add_graph(self, name: str, graphs: list["Row"], multiple=1):
        labels = [obj.name for obj in graphs]
        values = [obj.value * multiple for obj in graphs]
        self.graphs[name] = self.make_pie(values, labels)
        self.row_groups[name] = graphs

    @staticmethod
    def make_pie(values: list[Decimal], labels: list[str]) -> PieDict:
        return {
            'vals': [float(val) for val in values],
            'labels': [label.upper() for label in labels],
        }


class Row(object):

    def __init__(self, name: str, value: Decimal):
        self.name = name
        self.value = value

    def get_value(self) -> str:
        return str(self.value)

    def get_classes(self) -> str:
        return ''

    def __repr__(self):
        name = self.__class__.__name__
        return u'{}<{!r}, {!r}>'.format(name, self.name, self.value)


class Money(Row):

    def get_value(self) -> str:
        return u'{:0.2f} zł'.format(self.value)

    def get_classes(self) -> str:
        if self.value < 0:
            return 'negative'
        return ''


@dataclass
class PlotValues:
    name: str
    color: str
    moneys: list[float]


@dataclass
class PlotData:
    dates: list[str]
    moneys: list[float]
    acc: list[float]
    budgets: list[PlotValues]
    payment_types: list[PlotValues]


def get_costs_plot_data(day):
    KEY_FIELDS = (year_field, month_field)

    query_total = (
        db.session
        .query(
            sql_func.sum(Transaction.cost),
            *KEY_FIELDS,
        )
        .filter(Transaction.date >= day)
        .group_by(*KEY_FIELDS)
        .order_by(*KEY_FIELDS)
    )
    data = query_total.all()

    def format_key(o):
        return '{}-{:02}'.format(o[1], o[2])

    labels = [format_key(o) for o in data]

    def values_sorted_by_date(query):
        data = {format_key(o): o[0] for o in query.all()}
        return [float(data.get(key, 0)) for key in labels]

    def values_on_budget_query(budget) -> PlotValues:
        query = (
            db.session
            .query(
                sql_func.sum(Payment.cost),
                *KEY_FIELDS,
            )
            .join(Transaction)
            .filter(
                Transaction.date >= day,
                Payment.budget == budget,
            )
            .group_by(*KEY_FIELDS)
            .order_by(*KEY_FIELDS)
        )
        return PlotValues(
            name=budget.name,
            color='#' + budget.color,
            moneys=values_sorted_by_date(query),
        )

    def values_on_type_query(payment_type):
        query = (
            db.session
            .query(
                sql_func.sum(Payment.cost),
                *KEY_FIELDS,
            )
            .join(Transaction)
            .filter(
                Transaction.date >= day,
                Payment.payment_type == payment_type,
            )
            .group_by(*KEY_FIELDS)
            .order_by(*KEY_FIELDS)
        )
        return PlotValues(
            name=payment_type.name,
            color='#' + payment_type.color,
            moneys=values_sorted_by_date(query),
        )

    start_value = float(
        db.session
        .query(sql_func.sum(Transaction.cost))
        .filter(Transaction.date < day)
        .scalar() or 0.0
    )

    values = values_sorted_by_date(query_total)
    acc_values = [x + start_value for x in accumulate(values)]

    return PlotData(
        dates=labels,
        moneys=values,
        acc=acc_values,
        budgets=[
            values_on_budget_query(budget)
            for budget in (
                db.session
                .query(Budget)
                .filter(Budget.show_count_in_report == True)
            )
        ],
        payment_types=[
            values_on_type_query(payment_type)
            for payment_type in (
                db.session
                .query(PaymentType)
                .filter(PaymentType.show_count_in_report == True)
            )
        ]
    )


def get_data():
    day = date.today() - timedelta(days=365 * 1)
    day = day.replace(day=1)
    query_dates = db.session.query(
        year_field,
        month_field,
        sql_func.sum(Transaction.cost),
    ).group_by(
        year_field, month_field,
    ).order_by(
        year_field.desc(), month_field.desc(),
    ).filter(Transaction.date >= day)

    months = [
        ReportMonth(year, month, total)
        for year, month, total in query_dates.all()
    ]
    main_plot = get_costs_plot_data(day)

    return main_plot, months


@report_page.route(URL_PREFIX + '/report')
def basic():
    main_plot, months = get_data()

    return render_template(
        'report.html',
        months=months,
        main_plot=main_plot
    )


def make_workbook() -> Workbook:
    main_plot, months = get_data()
    wb = Workbook()

    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    title_style = NamedStyle("f:title")
    title_style.font = Font(name="Arial", bold=True, size=14)

    header_style = NamedStyle("f:header")
    header_style.font = Font(name="Arial", bold=True)
    header_style.border = border

    value_style = NamedStyle("f:value")
    value_style.number_format = "0.00 zł"
    value_style.font = Font(name="Arial")
    value_style.border = border

    wb.add_named_style(title_style)
    wb.add_named_style(header_style)
    wb.add_named_style(value_style)
    setup_styles(wb)

    now_year = datetime.now().year

    ws = wb.active
    assert ws
    ws.title = "Wykres"
    worksheet_plot(ws, plot_data=main_plot)
    worksheet_months(wb.create_sheet("Podsumowania per miesiąc"), months)

    query = (
        db.session.query(PaymentType.id, PaymentType.name)
        .filter(PaymentType.has_members)
    )
    for pt_id, pt_name in query:
        cebulany.resources.excels.paid_month.fill_worksheet(
            wb.create_sheet(pt_name),
            **cebulany.resources.excels.paid_month.get_data(now_year - 1, now_year, pt_id),
        )
    return wb


def cycle_color(colors: list):
    patterns = [None, 'zigZag', 'upDiag', 'dnDiag', 'smGrid']

    for p in patterns:
        for c in colors:
            pt = PatternFillProperties(prst=p)
            pt.foreground = ColorChoice(srgbClr=c)
            if not p:
                pt.background = ColorChoice(srgbClr=c)
            
            yield pt

positive_colors = ['63B76C', '7ED4E6', '7070CC', '4D8C57', 'F7A38E', 'E6BC5C', 'AF1F65']
negative_colors = ['F653A6', 'F2BA49', 'FD0E35', 'A9B2C3', '87421F', '8B8680', 'E6BE8A', 'D982B5']


def worksheet_months(ws, months: list[ReportMonth]):
    named_colors = {}
    cycle_colors = {
        'positive': cycle_color(positive_colors + negative_colors),
        'negative': cycle_color(negative_colors + positive_colors),
        'others': cycle_color(['FF0000']),
    }

    it_row = 1
    for report in months:
        it_header_row = it_row
        title = f"{report.year}-{report.month:02d}"
        title_cell = ws.cell(row=it_row, column=1, value=title)
        title_cell.style = "f:title"
        it_row += 1

        header_cell = ws.cell(row=it_row, column=1, value="Nazwa")
        header_cell.style = "f:header"
        header_cell = ws.cell(row=it_row, column=2, value="Wartosć")
        header_cell.style = "f:header"
        header_cell = ws.cell(row=it_row, column=3, value="ABS")
        header_cell.style = "f:header"
        it_row += 1

        color_rows = defaultdict(list)

        def add_cell(group: str, row: Row):
            nonlocal it_row
            key = row.name.upper()
            cell = ws.cell(row=it_row, column=1, value=key)
            cell.style = "f:header"
            cell = ws.cell(row=it_row, column=2, value=row.value)
            cell.style = "f:value"
            cell = ws.cell(row=it_row, column=3, value=abs(row.value))
            cell.style = "f:value"
            it_row += 1

            if key not in named_colors:
                named_colors[key] = next(cycle_colors[group]) 
            color_rows[group].append(named_colors[key])

        def add_graph(title, it, group_name, letter):
            end = it + len(report.row_groups[group_name]) - 1
            categories = Reference(ws, min_col=1, min_row=it, max_row=end)
            data = Reference(ws, min_col=3, min_row=it, max_row=end)
            chart = PieChart()
            chart.title = title
            chart.add_data(data)
            chart.set_categories(categories)
            chart.width = 15
            chart.height = 5

            for idx, pattern in enumerate(color_rows[group_name]):
                pt = DataPoint(idx=idx)
                pt.graphicalProperties.pattFill = pattern #ColorChoice(prstClr=color)
                chart.series[0].data_points.append(pt)

            ws.add_chart(chart, f"{letter}{it_header_row}")


        it_positive = it_row 
        for row in report.row_groups["positive"]:
            add_cell("positive", row)

        it_negative = it_row
        for row in report.row_groups["negative"]:
            add_cell("negative", row)

        for row in report.row_groups["others"]:
            add_cell("others", row)

        it_row += 1  # Span.

        add_graph(f"{title} - Przychód", it_positive, "positive", "E")
        add_graph(f"{title} - Koszt", it_negative, "negative", "N")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20


def worksheet_plot(ws, plot_data: PlotData):
    header_row = ["Miesiąc", "Zysk / Strata", "Saldo"]
    if plot_data.budgets:
        header_row += ["", "Miesiąc"] + [b.name for b in plot_data.budgets]
    if plot_data.payment_types:
        header_row += ["", "Miesiąc"] + [b.name for b in plot_data.payment_types]
    it_row_start = 20
    it_row = it_row_start

    for col, msg in enumerate(header_row, start=1):
        cell = ws.cell(row=it_row, column=col, value=msg)
        cell.style = "f:header"
        ws.column_dimensions[chr(65 + col)].width = 20
    it_row += 1

    it = zip(plot_data.dates, plot_data.moneys, plot_data.acc)

    for idx, (date, money, acc) in enumerate(it):
        cell = ws.cell(row=it_row, column=1, value=date)
        cell.style = "f:header"
        cell = ws.cell(row=it_row, column=2, value=money)
        cell.style = "f:value"
        cell = ws.cell(row=it_row, column=3, value=acc)
        cell.style = "f:value"
        col = 5

        if plot_data.budgets:
            cell = ws.cell(row=it_row, column=col, value=date)
            cell.style = "f:header"
            col += 1
            for budget in plot_data.budgets:
                cell = ws.cell(row=it_row, column=col, value=budget.moneys[idx])
                cell.style = "f:value"
                col += 1
            col += 1

        if plot_data.payment_types:
            cell = ws.cell(row=it_row, column=col, value=date)
            cell.style = "f:header"
            col += 1
            for payment_type in plot_data.payment_types:
                cell = ws.cell(row=it_row, column=col, value=payment_type.moneys[idx])
                cell.style = "f:value"
                col += 1
            col += 1

        it_row += 1

    rows = len(plot_data.moneys)

    def add_graph(cls, title, ref, min_col, max_col):
        data = Reference(ws, min_col=min_col + 1, min_row=it_row_start, max_col=max_col, max_row=it_row_start + rows)
        categories = Reference(ws, min_col=min_col, min_row=it_row_start + 1, max_row=it_row_start + rows)
        chart = cls()
        chart.title = title
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 25
        chart.height = 10
        chart.legend.legendPos = "b"
        chart_lines = ChartLines()
        chart_lines.graphicalProperties = GraphicalProperties(noFill=False)
        chart.x_axis.majorGridlines = chart_lines
        chart.y_axis.majorGridlines = chart_lines
        chart.y_axis.minorGridlines = chart_lines
        ws.add_chart(chart, ref)
        for s in chart.series:
            s.smooth = False
        return chart

    col = 1
    next_ref = "H1"

    main_plot = add_graph(AreaChart, "Finanse", ref="A1", min_col=col, max_col=col + 2)
    main_plot.series[0].graphicalProperties.line.solidFill = "009DC4"
    main_plot.series[0].graphicalProperties.line.width = 40000
    main_plot.series[0].graphicalProperties.noFill = True
    main_plot.series[1].graphicalProperties.line.noFill = True
    main_plot.series[1].graphicalProperties.solidFill = "76D7EA"
    # import pdb; pdb.set_trace()
    col += 4

    if plot_data.budgets:
        add_graph(LineChart, "Budżety", ref=next_ref, min_col=col, max_col=col + len(plot_data.budgets))
        col += len(plot_data.budgets) + 1
        next_ref = "O1"

    if plot_data.payment_types:
        add_graph(LineChart, "Typy płatności", ref=next_ref, min_col=col, max_col=col + len(plot_data.payment_types))
        col += len(plot_data.payment_types) + 1
