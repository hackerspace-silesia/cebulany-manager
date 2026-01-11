import calendar
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from functools import partial
from itertools import chain
from typing import Iterable, NamedTuple

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText

from cebulany.auth import token_required
from cebulany.models import Document, Payment, Transaction
from cebulany.queries.document import DocumentQuery
from cebulany.resources.excels.utils import color_text, send_excel, setup_styles, add_cell
from cebulany.resources.excels.blueprint import excel_page, URL_PREFIX
from cebulany.queries.transaction import TransactionQuery


@excel_page.route(URL_PREFIX + "/transactions/<int:year>-<int:month>")
@token_required
def excel_transaction(year: int, month: int):
    dt = f"{year}-{month:02d}"
    start = datetime(year, month, 1)
    end = datetime(year, month, calendar.monthrange(year, month)[1])
    transactions = TransactionQuery.get_transactions(date_range=(start, end))
    documents = DocumentQuery.get_documents(parent=dt).all()
    workbook = gen_workbook(dt, transactions, documents)
    download_name = f"transaction-{dt}.xlsx"
    return send_excel(workbook, download_name)


def gen_workbook(dt: str, transactions: list[Transaction], documents: Iterable[Document]):
    workbook = Workbook()
    setup_styles(workbook, font_size=8)
    fill_worksheet(workbook.active, dt, transactions)
    fill_attachments(workbook.create_sheet("Załączniki"), documents)

    return workbook


def fill_attachments(sheet, documents: Iterable[Document]):
    add = partial(add_cell, sheet)
    sheet.append(
        [
            add("Nazwa pliku", "header"),
            add("Numer", "header"),
            add("Podmiot", "header"),
            add("Data", "header"),
            add("Kwota", "header"),
            add("Komentarz", "header"),
        ]
    )

    sheet.column_dimensions["A"].width = 40.0
    sheet.column_dimensions["B"].width = 30.0
    sheet.column_dimensions["C"].width = 40.0
    sheet.column_dimensions["D"].width = 10.0
    sheet.column_dimensions["E"].width = 10.0
    sheet.column_dimensions["F"].width = 50.0

    for document in documents:
        sheet.append([
            add(document.filename, "left_header"),
            add(document.accounting_record, "wrap_text"),
            add(document.company_name, "wrap_text"),
            add(document.accounting_date and document.accounting_date.strftime("%Y-%m-%d"), "wrap_text"),
            add(document.price, "ok"),
            add(document.description, "wrap_text"),
        ])


def fill_worksheet(sheet, dt: str, transactions: list[Transaction]):
    sheet.title = f"Zestawienie {dt}"
    add = partial(add_cell, sheet)
    sheet.append(["", add(sheet.title, "header")])
    sheet.append(
        [
            add("L.p", "header"),
            add("Data", "header"),
            add("Kontrahent / Numer rachunku", "header"),
            add("Opis / Typ transakcji", "header"),
            add("Kwota", "header"),
            add("Numer dokumentu", "header"),
            add("Rodzaj", "header"),
            add("Źródło finansowania wydatku", "header"),
        ]
    )

    sheet.merge_cells(
        start_row=1,
        end_row=1,
        start_column=2,
        end_column=7,
    )

    sheet.column_dimensions["A"].width = 5.0
    sheet.column_dimensions["B"].width = 10.0
    sheet.column_dimensions["C"].width = 40.0
    sheet.column_dimensions["D"].width = 40.0
    sheet.column_dimensions["E"].width = 15.0
    sheet.column_dimensions["F"].width = 30.0
    sheet.column_dimensions["G"].width = 30.0
    sheet.column_dimensions["H"].width = 40.0
    sheet.freeze_panes = "B1"

    for index, transaction in enumerate(transactions, start=1):
        ps = transaction.payments
        payment_type = _to_rich_text(_format_payment_type(p) for p in ps)
        budget = _to_rich_text(_format_budget(p) for p in ps)
        if transaction.additional_info:
            budget.insert(0, "\n")
            budget.insert(0, transaction.additional_info)

        sheet.append(
            [
                add(f"{index:03d}", "left_header"),
                add(transaction.date.strftime("%Y-%m-%d"), "left_header"),
                add(transaction.name, "wrap_text"),
                add(transaction.title, "wrap_text"),
                add(transaction.cost, "bad" if transaction.cost < 0 else "nice"),
                add(_get_attachments(transaction), "wrap_text"),
                add(payment_type, "wrap_text_center"),
                add(budget, "wrap_text_center"),
            ]
        )

    for index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[index].height = 40.0


def _get_attachments(transaction: Transaction):
    atts = list(transaction.attachments)
    
    def cb(doc: Document):
        s = []
        if doc.company_name:
            s.append(doc.company_name)
        if doc.accounting_record:
            s.append(doc.accounting_record)
        return " ".join(s)

    return "; ".join(cb(att.document) for att in atts)


class RichCostLabel(NamedTuple):
    desc: str
    color: str
    cost: Decimal


def _format_payment_type(payment: Payment):
    pt = payment.payment_type
    at = pt.accountancy_type
    if at:
        desc = at.name
        color = at.color
    else:
        desc = pt.name
        color = pt.color

    return RichCostLabel(desc, color, abs(payment.cost))


def _format_budget(payment: Payment):
    if payment.cost >= 0:
        desc_attr = "description_on_positive"
    else:
        desc_attr = "description_on_negative"

    desc = getattr(payment.budget, desc_attr)
    color = payment.budget.color
    
    if payment.inner_budget:
        if additional_desc := getattr(payment.inner_budget, desc_attr):
            desc += f" {additional_desc}"
            color = payment.inner_budget.color

    return RichCostLabel(desc, color, abs(payment.cost))


def _to_rich_text(labels: Iterable[RichCostLabel]) -> CellRichText:
    accumulator: defaultdict[str, Decimal] = defaultdict(Decimal)
    colors: dict[str, str] = {}
    for label, color, cost in labels:
        accumulator[label] += cost
        colors[label] = color

    match len(accumulator):
        case 0:
            return CellRichText(["-"])
        case 1:
            label = next(iter(accumulator))
            color = colors[label]
            return CellRichText([color_text(label, color)])
        case _:
            elements = list(chain.from_iterable(
                [
                    color_text(label, colors[label]),
                    color_text(f" ({str(cost).replace('.', ',')} zł)", colors[label], bold=True),
                    "\n"
                ]
                for label, cost in accumulator.items()
            ))
            elements.pop()  # Remove last '\n'
            return CellRichText(elements)
