import re

from openpyxl import Workbook

from cebulany.auth import token_required
from cebulany.models import (
    db,
    PaymentType,
    Budget,
    InnerBudget,
)
from cebulany.resources.payment import query_parser
from cebulany.resources.excels.utils import (
    send_excel,
    setup_styles,
)
from cebulany.resources.excels.blueprint import excel_page, URL_PREFIX


from .common import ArgsObj
from .payments_worksheet import fill_payment_worksheet
from .payment_type_worksheet import fill_payment_type_worksheet
from .budget_worksheet import fill_budget_worksheet
from .inner_budget_worksheet import fill_inner_budget_worksheet



@excel_page.route(URL_PREFIX + "/payment/")
@token_required
def excel_payment():
    args = query_parser.parse_args()
    args.pop("page", None)

    obj_args = _get_args_obj(args)
    workbook = Workbook()
    setup_styles(workbook, font_size=8)

    workbook.active.title = "Zestawienie"
    fill_payment_worksheet(workbook.active, args, obj_args)

    if not obj_args.payment_type:
        fill_payment_type_worksheet(workbook.create_sheet("Rodzaj płatności"), args, obj_args)
    if not obj_args.budget:
        fill_budget_worksheet(workbook.create_sheet("Budżet"), args)
    if not obj_args.inner_budget:
        fill_inner_budget_worksheet(workbook.create_sheet("Budżet Wewnętrzny"), args)

    return send_excel(workbook, _get_download_name(obj_args))


def _get_args_obj(args):
    return ArgsObj(
        start_date=args["start_date"].date(),
        end_date=args["end_date"].date(),
        payment_type=db.session.execute(
            db.select(PaymentType).filter_by(id=args.payment_type_id)
        ).scalar_one_or_none(),
        budget=db.session.execute(
            db.select(Budget).filter_by(id=args.budget_id)
        ).scalar_one_or_none(),
        inner_budget=db.session.execute(
            db.select(InnerBudget).filter_by(id=args.inner_budget_id)
        ).scalar_one_or_none(),
    )


def _get_download_name(args: ArgsObj):
    start = args.start_date.strftime("%Y-%m-%d")
    end = args.end_date.strftime("%Y-%m-%d")
    name = "payments"

    if args.payment_type:
        name += f"_{args.payment_type.name}"
    if args.budget:
        name += f"_{args.budget.name}"
    if args.inner_budget:
        name += f"_{args.inner_budget.name}"

    return re.sub(r"\s+", "_", f"{name}_{start}_{end}.xlsx".replace(" ", "_"))



