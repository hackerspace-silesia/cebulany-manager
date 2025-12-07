from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from functools import partial
from itertools import groupby
from typing import Iterable, Protocol

from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText

from cebulany.models import (
    InnerTransfer,
    db,
    Payment,
    PaymentType,
    Budget,
    InnerBudget,
    Transaction,
)
from cebulany.resources.excels.utils import add_cell
from cebulany.queries.payment import PaymentQuery

from .common import ArgsObj, add_type_cell, make_inner_transfer_cell


def fill_payment_worksheet(sheet, args, obj_args: ArgsObj):
    factories: list[RowFactory] = [
        PaymentRowFactory(obj_args, key, list(group))
        for key, group in groupby(
            (
                db.session.execute(
                    PaymentQuery
                    .get_query_list(**args)
                    .order_by(None)
                    .order_by(*_make_payments_order(obj_args))
                ).scalars()
            ),
            key=lambda payment: (
                payment.payment_type,
                payment.budget,
                payment.inner_budget,
            ),
        )
    ]
    factories.append(
        InnerTransferRowFactory(
            obj_args,
            transfers=list(db.session.execute(
                PaymentQuery.get_inner_transfers(**args)
            ).scalars()),
        )
    )
    _fill_payment_worksheet(sheet, factories, obj_args)


def _make_payments_order(obj_args: ArgsObj):
    payments_order_by = []
    if not obj_args.payment_type:
        payments_order_by.append(Payment.payment_type_id)
    if not obj_args.budget:
        payments_order_by.append(Payment.budget_id)
    if not obj_args.inner_budget:
        payments_order_by.append(Payment.inner_budget_id)
    return payments_order_by + [Transaction.date.desc(), Transaction.line_num]


@dataclass
class Row:
    date: date | None
    name: str | CellRichText
    title: str | CellRichText
    payment_type: PaymentType | None
    budget: Budget | None
    inner_budget: InnerBudget | None
    cost: Decimal | None = None
    pos_cost: Decimal | None = None
    neg_cost: Decimal | None = None
    style: str = "ok"

    def make(self, sheet, args: ArgsObj, index: int = 0) -> list:
        row = [
            add_cell(sheet, f"{index:03d}" if index else "", "left_header"),
            add_cell(
                sheet,
                self.date.strftime("%Y-%m-%d") if self.date else "",
                "left_header",
            ),
            add_cell(sheet, self.name, self.style),
        ]

        if not args.payment_type:
            row.append(add_type_cell(sheet, self.payment_type))
        if not args.budget:
            row.append(add_type_cell(sheet, self.budget))
        if not args.inner_budget:
            row.append(add_type_cell(sheet, self.inner_budget))

        if self.pos_cost is None or self.neg_cost is None:
            assert self.cost is not None
            pos_cost = self.cost
            neg_cost = self.cost
        else:
            pos_cost = self.pos_cost
            neg_cost = self.neg_cost

        row += [
            add_cell(sheet, self.title, "wrap_text"),
            (
                add_cell(sheet, pos_cost, "nice")
                if pos_cost > 0
                else add_cell(sheet, "-", "wrap_text_center")
            ),
            (
                add_cell(sheet, neg_cost, "bad")
                if neg_cost < 0
                else add_cell(sheet, "-", "wrap_text_center")
            ),
        ]
        return row


class RowFactory(Protocol):
    def should_generate(self) -> bool: ...
    def make_title(self) -> str: ...
    def make_rows(self) -> Iterable[tuple[Row, Decimal]]: ...


class PaymentRowFactory(RowFactory):

    def __init__(self, args: ArgsObj, key, group: list[Payment]):
        elements = set()
        payment_type, budget, inner_budget = key
        if not args.payment_type:
            elements.add(payment_type.name.upper())
        if not args.budget:
            elements.add(budget.name.upper())
        if not args.inner_budget:
            elements.add(inner_budget.name.upper())
        self._group_name = "; ".join(elements)
        self._group = group

    def should_generate(self) -> bool:
        return bool(self._group)

    def make_title(self) -> str:
        return self._group_name

    def make_rows(self) -> Iterable[tuple[Row, Decimal]]:
        for payment in self._group:
            transaction: Transaction = payment.transaction
            description = transaction.title
            if transaction.additional_info:
                description = f"{transaction.additional_info}\n{description}"

            cost = payment.cost
            row = Row(
                date=transaction.date,
                name=payment.member.name if payment.member else payment.name,
                title=description,
                payment_type=payment.payment_type,
                budget=payment.budget,
                inner_budget=payment.inner_budget,
                cost=cost,
            )
            yield row, cost


class InnerTransferRowFactory:

    def __init__(self, args: ArgsObj, transfers: list[InnerTransfer]):
        self._budget = args.budget
        self._inner_budget = args.inner_budget
        self._transfers = transfers

    def should_generate(self) -> bool:
        return bool(self._transfers)

    def make_title(self) -> str:
        return "TRANSFER"

    def make_rows(self) -> Iterable[tuple[Row, Decimal]]:
        for transfer in self._transfers:
            title_cell, cost = make_inner_transfer_cell(
                transfer, self._budget, self._inner_budget
            )
            row = Row(
                date=transfer.date,
                name="Transfer",
                title=title_cell,
                payment_type=None,
                budget=transfer.budget,
                inner_budget=None,
                cost=cost,
            )
            yield row, cost


def _get_title(args):
    start = args.start_date.strftime("%Y-%m-%d")
    end = args.end_date.strftime("%Y-%m-%d")
    elements = []

    if args.payment_type:
        elements.append(args.payment_type.name)
    if args.budget:
        elements.append(args.budget.name)
    if args.inner_budget:
        elements.append(args.inner_budget.name)

    elements.append(f"{start} - {end}")
    return " ".join(elements)


def _fill_payment_worksheet(sheet, factories: Iterable[RowFactory], args: ArgsObj):
    add = partial(add_cell, sheet)
    header = _make_header(sheet, args)
    total_pos = Decimal("0.00")
    total_neg = Decimal("0.00")
    index = 0

    for factory in factories:
        if not factory.should_generate():
            continue
        sheet.append(["", add(factory.make_title(), "header")])
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=2,
            end_column=len(header),
        )

        factory_pos = Decimal("0.00")
        factory_neg = Decimal("0.00")
        for row, cost in factory.make_rows():
            index += 1
            if cost > 0:
                factory_pos += cost
            else:
                factory_neg += cost
            sheet.append(row.make(sheet, args, index))

        total_pos += factory_pos
        total_neg += factory_neg

        _add_cost_row(sheet, args, factory_pos, factory_neg)
        sheet.append([])


    sheet.append(["", add("PODSUMOWANIE", "header")])
    sheet.merge_cells(
        start_row=sheet.max_row,
        end_row=sheet.max_row,
        start_column=2,
        end_column=len(header),
    )
    _add_cost_row(sheet, args, total_pos, total_neg)

    for index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[index].height = 30.0


def _add_cost_row(sheet, args, pos: Decimal, neg: Decimal):
    first_sum_row = Row(
        date=None,
        name="RAZEM",
        title="",
        payment_type=None,
        budget=None,
        inner_budget=None,
        pos_cost=pos,
        neg_cost=neg,
        style="header",
    )
    sheet.append(first_sum_row.make(sheet, args))
    
    if pos and neg:
        second_sum_row = Row(
            date=None,
            name="RAZEM (suma)",
            title="",
            payment_type=None,
            budget=None,
            inner_budget=None,
            cost=pos + neg,
            style="header",
        )
        sheet.append(second_sum_row.make(sheet, args))


def _make_header(sheet, args):
    add = partial(add_cell, sheet)
    sheet.freeze_panes = "B3"

    header = [
        (add("Lp.", "header"), 5),
        (add("Data", "header"), 10),
        (add("Nazwa", "header"), 30),
    ]
    if not args.payment_type:
        header.append((add("Rodzaj", "header"), 15))
    if not args.budget:
        header.append((add("Budżet", "header"), 15))
    if not args.inner_budget:
        header.append((add("Budżet wew", "header"), 15))

    header += [
        (add("Tytuł", "header"), 50),
        (add("Przychód", "header"), 10),
        (add("Koszt", "header"), 10),
    ]

    sheet.append(["", add(f"Zestawienie {_get_title(args)}", "header")])
    sheet.merge_cells(
        start_row=1,
        end_row=1,
        start_column=2,
        end_column=len(header),
    )

    sheet.append([col for col, _ in header])

    for column_count, (_, width) in enumerate(header, start=1):
        letter = get_column_letter(column_count)
        sheet.column_dimensions[letter].width = width


    return header
