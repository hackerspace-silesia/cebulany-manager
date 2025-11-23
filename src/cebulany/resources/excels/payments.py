from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from functools import partial
from typing import Iterable, Protocol
import re

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText

from cebulany.auth import token_required
from cebulany.models import (
    InnerTransfer,
    db,
    Payment,
    PaymentType,
    Budget,
    InnerBudget,
    Transaction,
)
from cebulany.queries.inner_transfer import InnerTransferQuery
from cebulany.resources.payment import query_parser
from cebulany.resources.excels.utils import (
    color_text,
    send_excel,
    setup_styles,
    add_cell,
)
from cebulany.resources.excels.blueprint import excel_page, URL_PREFIX
from cebulany.queries.payment import PaymentQuery


@dataclass
class ArgsObj:
    start_date: date
    end_date: date
    payment_type: PaymentType | None
    budget: Budget | None
    inner_budget: InnerBudget | None


class TypeLike(Protocol):
    name: str
    color: str
    cost: Decimal


@excel_page.route(URL_PREFIX + "/payment/")
@token_required
def excel_payment():
    args = query_parser.parse_args()
    args.pop("page", None)

    obj_args = _get_args_obj(args)
    workbook = Workbook()
    setup_styles(workbook, font_size=8)

    payments = list(db.session.execute(PaymentQuery.get_query_list(**args)).scalars())
    payments += list(
        db.session.execute(PaymentQuery.get_inner_transfers(**args)).scalars()
    )
    payments.sort(key=_sort_key)
    sum_query = PaymentQuery.get_query_agg(**args)
    _, total = db.session.execute(sum_query).one()
    _fill_payment_worksheet(workbook.active, payments, obj_args, total)

    if not obj_args.payment_type:
        transfers = InnerTransferQuery.get_agg_query(
            start_date=args.start_date,
            end_date=args.end_date,
            inner_budget_id=args.inner_budget_id,
            budget_id=args.budget_id,
        )
        payment_types = _enhance_query(
            PaymentQuery.get_query_group_by_type(**args), PaymentType
        )
        sheet = workbook.create_sheet("Rodzaj płatności")
        _fill_type_worksheet(
            sheet,
            _types_to_rows(sheet, payment_types),
            _inner_transfers_to_rows(
                sheet,
                obj_args.budget,
                obj_args.inner_budget,
                db.session.execute(transfers),
            ),
        )
    if not obj_args.budget:
        sheet = workbook.create_sheet("Budżet")
        _fill_type_worksheet(
            sheet,
            _types_to_rows(
                sheet,
                _enhance_query(
                    PaymentQuery.get_query_group_by_budget(**args).join(Budget), Budget
                ),
            ),
        )
    if not obj_args.inner_budget:
        sheet = workbook.create_sheet("Budżet Wewnętrzny")
        _fill_type_worksheet(
            sheet,
            _types_to_rows(
                sheet,
                _enhance_query(
                    PaymentQuery.get_query_group_by_inner_budget(**args).join(
                        InnerBudget
                    ),
                    InnerBudget,
                ),
            ),
        )

    return send_excel(workbook, _get_download_name(obj_args))


def _sort_key(o):
    if isinstance(o, InnerTransfer):
        return (o.date, 0)
    else:
        return (o.transaction.date, o.transaction.line_num)


def _enhance_query(query, model) -> Iterable[TypeLike]:
    name = model.name.label("name")
    color = model.color.label("color")
    query = query.add_columns(name, color).group_by(name, color)
    return db.session.execute(query)


def _types_to_rows(sheet, query: Iterable[TypeLike]):
    for obj in query:
        yield [
            add_type_cell(sheet, obj),
            add_cell(sheet, obj.cost, "bad" if obj.cost < 0 else "nice"),
        ]


def _inner_transfers_to_rows(
    sheet,
    budget: Budget | None,
    inner_budget: InnerBudget | None,
    transfers: Iterable[InnerTransfer],
):
    for obj in transfers:
        title_cell, cost = make_inner_transfer_cell(
            obj, budget, inner_budget, "Transfer"
        )
        yield [
            add_cell(sheet, title_cell, "wrap_text_center"),
            add_cell(sheet, cost, "bad" if cost < 0 else "nice"),
        ]


def _get_args_obj(args):
    return ArgsObj(
        start_date=args["start_date"].date(),
        end_date=args["end_date"].date(),
        payment_type=db.session.execute(
            db.select(PaymentType).filter_by(id=args.payment_type_id)
        ).scalar_one_or_none(),
        budget=db.session.execute(
            db.select(Budget).filter_by(id=args.budget_id)
        ).scalar_one_or_none(),
        inner_budget=db.session.execute(
            db.select(InnerBudget).filter_by(id=args.inner_budget_id)
        ).scalar_one_or_none(),
    )


def _get_download_name(args: ArgsObj):
    start = args.start_date.strftime("%Y-%m-%d")
    end = args.end_date.strftime("%Y-%m-%d")
    name = "payments"

    if args.payment_type:
        name += f"_{args.payment_type.name}"
    if args.budget:
        name += f"_{args.budget.name}"
    if args.inner_budget:
        name += f"_{args.inner_budget.name}"

    return re.sub(r"\s+", "_", f"{name}_{start}_{end}.xlsx".replace(" ", "_"))


def add_type_cell(sheet, type: TypeLike | str | None):
    if isinstance(type, str):
        return add_cell(sheet, type, "wrap_text_center")
    if type is None:
        return add_cell(sheet, "", "wrap_text_center")
    cell = add_cell(sheet, type.name or "-", "wrap_text_center")
    cell.font = Font(name="Calibri", bold=True, size=8, color=type.color)
    return cell


def add_type_text_block(type: TypeLike | None, default: str = "-"):
    if type is None:
        return color_text(default, bold=True)
    return color_text(type.name, type.color, bold=True)


def make_inner_transfer_cell(transfer: InnerTransfer, budget, inner_budget, prefix=""):
    title_cell = CellRichText(prefix) if prefix else CellRichText()
    factor = 1
    inner_budget_id = inner_budget.id if inner_budget else None
    if transfer.from_id != inner_budget_id:
        if title_cell:
            title_cell.append(" ")
        title_cell += [
            "z ",
            add_type_text_block(transfer.from_inner_budget),
        ]
    if transfer.to_id != inner_budget_id:
        factor = -1
        if title_cell:
            title_cell.append(" ")
        title_cell += [
            "do ",
            add_type_text_block(transfer.to_inner_budget),
        ]
    budget_id = budget.id if budget else None
    if transfer.budget_id != budget_id:
        if title_cell:
            title_cell.append(" ")
        title_cell += [
            "(",
            add_type_text_block(transfer.budget),
            ")",
        ]

    return title_cell, transfer.cost * factor


def _fill_payment_worksheet(
    sheet, payments: Iterable[Payment | InnerTransfer], args: ArgsObj, total: Decimal
):
    add = partial(add_cell, sheet)
    sheet.title = "Zestawienie"
    sheet.append(["", add(f"Zestawienie {_get_title(args)}", "header")])
    total_pos = Decimal("0.00")
    total_neg = Decimal("0.00")

    def make_row(
        index: int | None,
        date: date | None,
        name: str | CellRichText,
        title: str | CellRichText,
        payment_type,
        budget,
        inner_budget,
        cost: Decimal | None = None,
        pos_cost: Decimal | None = None,
        neg_cost: Decimal | None = None,
        style=None,
    ):
        row = [
            add(f"{index:03d}" if index else "", "left_header"),
            add(date.strftime("%Y-%m-%d") if date else "", "left_header"),
            add(name),
        ]

        if not args.payment_type:
            row.append(add_type_cell(sheet, payment_type))
        if not args.budget:
            row.append(add_type_cell(sheet, budget))
        if not args.inner_budget:
            row.append(add_type_cell(sheet, inner_budget))

        if pos_cost is None or neg_cost is None:
            assert cost is not None
            pos_cost = cost
            neg_cost = cost

        row += [
            add(title, "wrap_text"),
            add(pos_cost, "nice") if pos_cost > 0 else add("-", "wrap_text_center"),
            add(neg_cost, "bad") if neg_cost < 0 else add("-", "wrap_text_center"),
        ]
        return row

    header = [
        (add("Lp.", "header"), 5),
        (add("Data", "header"), 10),
        (add("Nazwa", "header"), 30),
    ]
    if not args.payment_type:
        header.append((add("Rodzaj", "header"), 15))
    if not args.budget:
        header.append((add("Budżet", "header"), 15))
    if not args.inner_budget:
        header.append((add("Budżet wew", "header"), 15))

    header += [
        (add("Tytuł", "header"), 60),
        (add("Zysk", "header"), 15),
        (add("Koszt", "header"), 15),
    ]
    sheet.append([cell for cell, _ in header])

    sheet.merge_cells(
        start_row=1,
        end_row=1,
        start_column=2,
        end_column=len(header),
    )

    for column_count, (_, width) in enumerate(header, start=1):
        letter = get_column_letter(column_count)
        sheet.column_dimensions[letter].width = width

    sheet.freeze_panes = "B1"

    for index, payment in enumerate(payments, start=1):
        if isinstance(payment, Payment):
            transaction: Transaction = payment.transaction
            description = transaction.title
            if transaction.additional_info:
                description = f"{transaction.additional_info}\n{description}"

            cost = payment.cost
            row = make_row(
                index=index,
                date=transaction.date,
                name=payment.member.name if payment.member else payment.name,
                title=description,
                payment_type=payment.payment_type,
                budget=payment.budget,
                inner_budget=payment.inner_budget,
                cost=cost,
            )
        elif isinstance(payment, InnerTransfer):
            title_cell, cost = make_inner_transfer_cell(
                payment, args.budget, args.inner_budget
            )
            row = make_row(
                index=index,
                date=payment.date,
                name="Transfer",
                title=title_cell,
                payment_type=None,
                budget=payment.budget,
                inner_budget=None,
                cost=cost,
            )
        else:
            row = [add("???", "wrap_text")]
            cost = Decimal(0)

        if cost > 0:
            total_pos += cost
        else:
            total_neg += cost
        sheet.append(row)

    first_sum_row = make_row(
        index=None,
        date=None,
        name="RAZEM",
        title="",
        payment_type=None,
        budget=None,
        inner_budget=None,
        pos_cost=total_pos,
        neg_cost=total_neg,
    )
    second_sum_row = make_row(
        index=None,
        date=None,
        name="RAZEM (suma)",
        title="",
        payment_type=None,
        budget=None,
        inner_budget=None,
        cost=total,
    )
    sheet.append(first_sum_row)
    sheet.append(second_sum_row)

    for index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[index].height = 30.0


def _get_title(args):
    start = args.start_date.strftime("%Y-%m-%d")
    end = args.end_date.strftime("%Y-%m-%d")
    name = ""

    if args.payment_type:
        name += f" {args.payment_type.name}"
    if args.budget:
        name += f" {args.budget.name}"
    if args.inner_budget:
        name += f" {args.inner_budget.name}"

    return f"{name} {start} - {end}"


def _fill_type_worksheet(sheet, *objs: Iterable):
    add = partial(add_cell, sheet)

    sheet.append([add(sheet.title, "header")])
    sheet.merge_cells(
        start_row=1,
        end_row=1,
        start_column=1,
        end_column=2,
    )

    sheet.append(
        [
            add("Nazwa", "header"),
            add("Kwota", "header"),
        ]
    )

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 10

    for rows in objs:
        for row in rows:
            sheet.append(row)

    sheet.freeze_panes = "A2"
